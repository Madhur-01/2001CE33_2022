#MADHUR GARG 2001CE33
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment,Border,Side
import numpy as np
import os
import re
from datetime import datetime
os.system('cls')
start_time = datetime.now()

def OnE_iNnInG(inn1,bat_pl,bow_pl,s) : #making function for one innings
    innbat = Workbook()
    innfow = Workbook()
    innbow = Workbook()
    S1 = innbat.active
    S2 = innbow.active
    S3 = innfow.active
    S1.column_dimensions['A'].width = 25


    #scorecard and index
    S1['A1']    = s + ' Innings'
    S1['I1']    = '0-0'
    S1['J1']    = '0 overs'
    S1['A2']    = 'Batter'
    S1['F2']    = 'R'
    S1['G2']    = 'B'
    S1['H2']    = '4s'
    S1['I2']    = '6s'
    S1['J2']    = 'SR'


    S2['A1']    = 'Bowler'
    S2['D1']    =  'O'
    S2['E1']    = 'M'
    S2['F1']    = 'R'
    S2['G1']    =  'W'
    S2['H1']    = 'NB'
    S2['I1']    =  'WD'
    S2['J1']    = 'ECO'

    
    S3['A1']    = 'Fall of wickets'
    

    #using regex
    Over      =   re.compile(r'(\d\d?\.\d)')
    Zero      =   re.compile(r'no run')
    No_ball   =   re.compile(r', (no ball),')
    Wide      =   re.compile(r', wide,')
    Wide2     =   re.compile(r', 2 wides,')
    wide3     =   re.compile(r', 3 wides,')
    single    =   re.compile(r', 1 run,')
    SIX       =   re.compile(r', SIX,')
    FOUR      =   re.compile(r', FOUR,')
    BYES      =   re.compile(r', (byes),')
    lBYES     =   re.compile(r', (leg byes),')
    Double    =   re.compile(r', 2 runs,')
    Triple    =   re.compile(r', 3 runs,')
    out       =   re.compile(r', out')
    player    =   re.compile(r'(\d\d?\.\d) (\w+) (to|\w+ to) (\w+)( \w+)?,')
    caught    =   re.compile(r', out Caught by (\w+)')
    lbw       =   re.compile(r', out Lbw!!')
    BOWLED    =   re.compile(r', out bowled!!')
    run_out   =   re.compile(r'Run Out!! ')
    runs      =   0
    wickets   =   0
    nb        =   0
    nlb       =   0
    nw        =   0
    nnb       =   0
    ppr       =   0

 #coding dynamically based on each line
    while True:
        t = inn1.readline()
        if not t:
            break
        crun = 0
        cex = 0
        ov = Over.findall(t)
        if float(ov[0]) == int(float(ov[0])) + 0.6:
            ov[0] = str(int(float(ov[0]))+1)
        if float(ov[0])== int(float(ov[0])) + 0.1:
            orun = runs
        nnnb = No_ball.findall(t)
        wbb = Wide.findall(t)
        sr = single.findall(t)
        zr = Zero.findall(t)
        by = BYES.findall(t)
        lby = lBYES.findall(t)
        pl = player.finditer(t)
        w2 = Wide2.findall(t)
        w3 = wide3.findall(t)
        for i in pl:
            cbow = i.group(2)
            cbat = i.group(4)
        r = 1
        for col in S1.iter_cols(min_col=1,max_col = 1):
            for cell in col:
                if  cell.value == bat_pl[cbat]:
                    crow = cell.row
                    r = 0
                    break
        if r:
            S1.append([bat_pl[cbat],'not out','','','',0,0,0,0,0])
            crow = S1.max_row

        s = 1
        for col in S2.iter_cols(min_col=1,max_col = 1):
            for cell in col:
                if  cell.value == bow_pl[cbow]:
                    cbrow = cell.row
                    s = 0
                    break
        if s:
            S2.append([bow_pl[cbow],'','',0,0,0,0,0,0,0])
            cbrow = S2.max_row

        #making different cases for different rules in cricket
        if sr:
            crun = 1
        elif zr:
            pass
        else:
            db = Double.findall(t)
            if db:
                crun = 2
            else:
                fr = FOUR.findall(t)
                if fr:
                    crun = 4
                    if not (by or lby):
                        S1.cell(row = crow,column = 8).value = S1.cell(row = crow,column = 8).value + 1
                else:
                    sx = SIX.findall(t)
                    if sx:
                        crun = 6
                        S1.cell(row = crow,column = 9).value = S1.cell(row = crow,column = 9).value + 1
                    else:
                        tp = Triple.findall(t)
                        if tp:
                            crun = 3
        if wbb or nnnb or w2 or w3:
            cex = 1
            S2.cell(row = cbrow,column = 6).value = S2.cell(row = cbrow,column = 6).value + 1
            if wbb:
                nw = nw + 1
                S2.cell(row = cbrow,column = 9).value = S2.cell(row = cbrow,column = 9).value + 1 
            elif w2:
                nw = nw + 2
                cex = 2
                S2.cell(row = cbrow,column = 6).value = S2.cell(row = cbrow,column = 6).value + 1
                S2.cell(row = cbrow,column = 9).value = S2.cell(row = cbrow,column = 9).value + 2
            elif w3:
                nw = nw + 3
                cex = 3
                S2.cell(row = cbrow,column = 6).value = S2.cell(row = cbrow,column = 6).value + 2
                S2.cell(row = cbrow,column = 9).value = S2.cell(row = cbrow,column = 9).value + 3
            else:
                nnb = nnb + 1
                S1.cell(row = crow,column = 7).value = S1.cell(row = crow,column = 7).value + 1
                S2.cell(row = cbrow,column = 8).value = S2.cell(row = cbrow,column = 8).value + 1
        else:
            S1.cell(row = crow,column = 7).value = S1.cell(row = crow,column = 7).value + 1
            bo = 10*S2.cell(row = cbrow,column = 4).value -int(S2.cell(row = cbrow,column = 4).value)*4 + 1
            S2.cell(row = cbrow,column = 4).value = int(bo/6)*0.4 + bo*0.1
        runs = runs + crun + cex
        if float(ov[0]) < 6.1:
            ppr = runs
        if float(ov[0])== int(float(ov[0])):
            if orun == runs:
                S2.cell(row = cbrow,column = 5).value = S2.cell(row = cbrow,column = 5).value + 1
        ot = out.findall(t)
        ct = caught.findall(t)
        lw = lbw.findall(t)
        bw = BOWLED.findall(t)
        ctu = caught.finditer(t)
        ro = run_out.findall(t)
        for i in ctu:
            capl = i.group(1)
        if ot:
            wickets = wickets + 1
            if not ro:
                S2.cell(row = cbrow,column = 7).value = S2.cell(row = cbrow,column = 7).value + 1
            if wickets != 1:
                S3['A2'] = str(S3['A2'].value) + ', ' + str(runs) + '-' + str(wickets) + ' (' + bat_pl[cbat] + ', ' + ov[0] + ')'
            else:
                S3['A2'] = str(runs) + '-' + str(wickets) + '(' + bat_pl[cbat] + ', ' + ov[0] + ')'
            if ct:
                S1.cell(row = crow,column = 2).value = 'c ' + bow_pl[capl] + ' b ' + bow_pl[cbow]
            elif lw:
                S1.cell(row = crow,column = 2).value = 'lbw b ' + bow_pl[cbow]
            elif bw:
                S1.cell(row = crow,column = 2).value = 'b ' + bow_pl[cbow]
            elif ro:
                S1.cell(row = crow,column = 2).value = 'run out (' + bow_pl[cbow] + ')'
            
        #filling in excel
        S1['I1'] = str(runs)+'-'+str(wickets)+'('+ov[0]+' Ov)'
        
        if not (by or lby):
            S1.cell(row = crow,column = 6).value = S1.cell(row = crow,column = 6).value + crun
            S2.cell(row = cbrow,column = 6).value = S2.cell(row = cbrow,column = 6).value + crun
        else:
            if by:
                if sr:
                    nb = nb + 4
                elif db:
                    nb = nb + 2
                elif tp:
                    nb = nb + 3
                elif fr:
                    nb = nb + 1
            else:
                if sr:
                    nlb = nlb + 1
                elif fr:
                    nlb = nlb + 4
                elif db:
                    nlb = nlb + 2
                elif tp:
                    nlb = nlb + 3
            

        S1.cell(row = crow,column = 10).value = float("{:.2f}".format((S1.cell(row = crow,column = 6).value)*100/S1.cell(row = crow,column = 7).value))
        ovf = (10*S2.cell(row = cbrow,column = 4).value - 4*int(S2.cell(row = cbrow,column = 4).value))/6
        if ovf:
            S2.cell(row = cbrow,column = 10).value = float("{:.2f}".format(S2.cell(row = cbrow,column = 6).value/ovf))
        t = inn1.readline()
        if not t:
            break

    #formatting cells
    S1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    S1.merge_cells(start_row=1, start_column=9, end_row=1, end_column=10)
    crow = S1.max_row + 1
    S1.merge_cells(start_row=crow, start_column=5, end_row=crow, end_column=7)

    S1.append(['Extras','','','','','','',str(nb + nlb + nw + nnb)+'(b '+str(nb)+', lb '+str(nlb)+', w '+str(nw)+', nb '+str(nnb)+', p 0)'])
    S1.merge_cells(start_row=crow+1, start_column=5, end_row=crow+1, end_column=7)
    S1.append(['Total','','','','','','',str(runs)+'('+str(wickets)+' wkts, '+ov[0]+' Ov)'])
    
    S2.append([])
    S2.append(['Powerplays','Overs','','','','','','','Runs'])
    S2.append(['Mandotary','0.1-6','','','','','','',ppr])

    #combining data
    mc1 = S1.max_column
    mc2 = S2.max_column
    mc3 = S3.max_column
    mr1 = S1.max_row
    mr2 = S2.max_row
    mr3 = S3.max_row


    for i in range(1,mr1-3):
        S1.merge_cells(start_row=i+1, start_column=2, end_row=i+1, end_column=5)
    for i in range(1,mr3+1):
        for j in range(1,9):
            S1.cell(row = mr1+i+1,column=j).value = S3.cell(row = i,column=j).value

    mr11 = S1.max_row
    for i in range(1,mr2+1):
        for j in range(1,11):
            S1.cell(row = mr11+i+3,column=j).value = S2.cell(row = i,column=j).value

    #few more formatting
    S1.merge_cells(start_row=mr1 + 2, start_column=1, end_row=mr1 + 2, end_column=10)
    S1.merge_cells(start_row=mr1 + 3, start_column=1, end_row=mr1 + 5, end_column=10)
    S1.cell(row = mr1+3,column=1).alignment = Alignment(wrap_text=True)
    mr12 = S1.max_row
    S1.merge_cells(start_row=mr12 -1, start_column=2, end_row=mr12-1, end_column=8)
    S1.merge_cells(start_row=mr12, start_column=2, end_row=mr12, end_column=8)
    S1.merge_cells(start_row=mr12 -1, start_column=9, end_row=mr12-1, end_column=10)
    S1.merge_cells(start_row=mr12, start_column=9, end_row=mr12, end_column=10)

    return innbat

#function for socrecard generation
def ScOEcArD(inn1,inn2):
    pak_player = {'Babar':'Babar Azam (c)','Rizwan':'Mohammad Rizwan (wk)','Fakhar':'Fakhar Zaman','Iftikhar':'Iftikhar Ahmed','Khushdil':'Khushdil Shah','Asif':'Asif Ali','Shadab':'Shadab Khan','Mohammad':'Mohammad Nawaz','Naseem':'Naseem Shah','Haris':'Haris Rauf','Dahani':'Shahnawaz Dahani'}
    ind_player = {'Rohit':'Rohit Sharma (c)','Rahul':'KL Rahul','Kohli':'Virat Kohli','Suryakumar':'Suryakumar Yadav','Jadeja':'Ravindra Jadeja','Hardik':'Hardik Pandya','Karthik':'Dinesh Karthik (wk)','Bhuvneshwar':'Bhuvneshwar Kumar','Avesh':'Avesh Khan',"Arshdeep":'Arshdeep Singh','Chahal':'Yuzvendra Chahal'}
    finn = OnE_iNnInG(inn1,pak_player,ind_player,'Pakistan')
    sinn = OnE_iNnInG(inn2,ind_player,pak_player,'India')

    S1 = finn.active
    S2 = sinn.active

    #combining both the innings
    mc1 = S1.max_column
    mc2 = S2.max_column
    mr1 = S1.max_row
    mr2 = S2.max_row

    for i in range(1,mr2+1):
        for j in range(1,11):
            S1.cell(row = mr1+i+1,column=j).value = S2.cell(row = i,column=j).value    
    S1.merge_cells(start_row=34, start_column=1, end_row=34, end_column=8)
    S1.merge_cells(start_row=35, start_column=2, end_row=35, end_column=5)
    S1.merge_cells(start_row=48, start_column=1, end_row=50, end_column=10)
    for i in range(36,43):
        S1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=5)
    

    return finn
    
#opening files
pak = open("pak_inns1.txt",'r')
ind = open('india_innS2.txt','r')
team = open('teams.txt','r')

wb = Workbook()
wb = ScOEcArD(pak,ind)

#saving the output
wb.save('scorecard.xlsx')






end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
