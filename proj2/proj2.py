#Madhur: 2001CE33
#Siddharth: 2001CE59

#Importing all required libraries.
from datetime import datetime
start_time = datetime.now()

#Importing strimlit for front-end interface.
import streamlit as st
import pandas as pandas
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border, Side

#This will be used for creating dialogue box(First part of code-strimlit.)
import tkinter as tk
from tkinter import filedialog
import glob
import os
from zipfile import ZipFile

#Already given part of the code.
from platform import python_version

os.system("cls")


ver = python_version()

#Here we will be reading all the files (or selected one) from input folder. (strictly .xlsx)
#After that code from tut.7 will be working in the backend which will chnange the files for a desired output and a proper naming scheme is followed.
#Finally all the files will be downlaoded or saved throgh the interface developed through streamlit.

#These are some formating criterias and values for color and other things which will be used in for formating the cells and tables of output.
OcT_SigN = [1, -1, 2, -2, 3, -3, 4, -4]

OcT_NaMe_Id_MaPPinG = {1: "Internal outward interaction", -1: "External outward interaction", 2: "External Ejection", -2: "Internal Ejection", 3: "External inward interaction", -3: "Internal inward interaction", 4: "Internal sweep", -4: "External sweep"}

YeLLoW = "00FFFF00"

YeLLoW_bG = PatternFill(start_color=YeLLoW, end_color=YeLLoW, fill_type='solid')

Blc = "00000000"

DoUbLe = Side(border_style="thin", color=Blc)

Blc_BORDER = Border(top=DoUbLe, left=DoUbLe, right=DoUbLe, bottom=DoUbLe)

#This variable will be used to store the file uploaded on interface. 

NeW_fIlE=None

MoD=0 #Initialised MoD value to 0.

#Function for front-end part.
def StReAmLiT_PrOj2():

    #Setting the webpage title
    st.set_page_config(page_title="Team: Madhur and Siddharth")
    
    path="" #This will be storing the path of folder from where we will be uploading our files.
    
    st.header("Welcome, this is GUI for Project 2.")
    st.subheader("And this is team Madhur n Siddharth.")
    #Using the radio button, to give options for the type of conversion.
    #Setting the header value of the webpage

    #Using the radio button, to give options for the type of conversion.

    STATUS = st.radio("Select conversion type: ", ('Single file Conversion', 'Bulk file Conversion'))
    
    #On selecting Single file Conversion.

    if (STATUS == 'Single file Conversion'):
        global NeW_fIlE
        
        #Using file_uploader for uploading the file.

        NeW_fIlE = st.file_uploader("Upload the input dataset: ", type="xlsx")     #Strictly for .xlsx files only.
        if "path" in st.session_state: #Basically for re-running the program.
            del st.session_state["path"]
    
    #On selecting Bulk file Conversion.

    if (STATUS=='Bulk file Conversion'):
        #Here we will be starting by selecting the folder.
        NeW_fIlE=None
        
        #This will be used for bringing the dialogue box, where option for selecting input folder will be given.

        RooT = tk.Tk()
        RooT.withdraw()

        RooT.wm_attributes('-topmost', 1) #This will open the dialog box in top left part of screen.

		#Creating a button to pick folder.

        st.write('Please select a folder:')
        ClIcKeD = st.button('Choose the folder')

        if ClIcKeD:
            path = filedialog.askdirectory(master=RooT) #Through this the dialogue box will ask for directory.
            st.session_state["path"] = path

    #Brings back to gui window.
    if "path" in st.session_state:




        path = st.session_state["path"]
        dirname = st.text_input('Selected folder:', path) #Here the addresss of that folder is shown.
        
    global MoD
    #After selecting the input files, here we will collect the MoD value.
    MoD = st.number_input('Enter MoD value: ', min_value=1,  value=5000, step=1)

    #Devided the section in 2 columns, left side will have button for conversion, whereas right side will have button for downloading the file.
    CoNvErT, download=st.columns(2)
    
    #download(Wanna download) button will apppear only if, CoNvErT(Compute) button is used.
    with CoNvErT:
        CoNv_BuT=st.button("Compute") #Button for conversion.
        
        #After clicling on Compute.
        if CoNv_BuT:
            #Case for Single file Conversion.
            if (STATUS == 'Single file Conversion'): 
                
                #To assure that a file is selected.
                if not NeW_fIlE:
                    st.warning("Please upload a file!!")
                else:
                    
                    #Removing file extension cause we will be passing only file name to the main function(tut 7).
                    #Saving the nam in FiLe_NaMe.

                    FiLe_NaMe=NeW_fIlE.name.split(".xlsx")[0]
                    outputFileName=PROJ2(FiLe_NaMe) #Passing this name to main function.
                    
                    with download:
                        #Opening/recieving the output file, and dowloading it.
                        
                        with open(outputFileName, 'rb') as req_file:
                            st.download_button(label="Wanna download", data=req_file, file_name=outputFileName, mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
              
            #Case of Bulk file conversion.      
            elif (STATUS=='Bulk file Conversion'):
                if "path" in st.session_state:
                    path = st.session_state["path"]

                #Here we will be using the path of the folder and using a loop we will CoNvErT all the files.
                folder = path.split("/")[-1]

                EXCEL_FILES=glob.glob(os.path.join(path,"*.xlsx"))#All the files in input with .xlsx extension.
                
                #To assure that required input is selected.
                if(len(folder)==0):
                    st.warning("Atleast select a folder!!")
                    return

                if len(EXCEL_FILES)==0:
                    st.warning("No excel file present!!")
                    return 
                
                #Using the give namimg scheme for namimg the output folder.
                OuTpUtFoLdErNaMe=GeTNaMe(folder)+".zip"
                
                #ZipFile will zip the folder.
                ZiPoBJ=ZipFile(OuTpUtFoLdErNaMe,'w')
                
                #Using loop and moving through all the excel files prsent in input and CoNvErTing them to output.
                for i, file in enumerate(EXCEL_FILES):
                    NeW_fIlE=file
                    FiLe_NaMe=file.split(".xlsx")[0]
                    FiLe_NaMe=FiLe_NaMe.split("\\")[-1]
                    
                    #Calling main function for conversion of each file of excel files. 
                    outputFileName=PROJ2(FiLe_NaMe)
                    
                    #Adding the file to zip folder.
                    ZiPoBJ.write(outputFileName)
                    
                ZiPoBJ.close()
                
                #Finally downloading the zip folder.
                with download:
                    with open(OuTpUtFoLdErNaMe,'rb') as req_file:
                        st.download_button(label="Wanna Downalod", data=req_file, file_name=OuTpUtFoLdErNaMe)

#This function is used for implemention naming scheme.
def GeTNaMe(inputFileName):
    current=datetime.now()
    #Adding date n time.
    dt_s=current.strftime("%Y-%m-%d-%H-%M-%S")
    
    #Adding output/ in front for saving all the files in output folder after changing the names.
    outputFileName="output/"+inputFileName+"_MoD_"+str(MoD)+"_"+dt_s
    return outputFileName

#This is a MoDified version of tut 7 code, although all the function used inside this are same.
def PROJ2(FiLe_NaMe):
    DatAFraMe=pandas.read_excel(NeW_fIlE)
    outputFileName=GeTNaMe(FiLe_NaMe)+".xlsx"
    
    outputFile=openpyxl.Workbook()
    CURRENT_SHEET=outputFile.active
    
    outputFile = openpyxl.Workbook()
    CURRENT_SHEET = outputFile.active
    ToT_CoUnT = 0

    col = 1

    #Declairing variables to store sum values.
    U_sum = 0 
    V_sum = 0
    W_sum = 0

    for key, value in DatAFraMe.items():
        value = value.tolist()
        ToT_CoUnT = len(value)

        #Key shifted to 2nd row.
        CURRENT_SHEET.cell(row=2, column=col).value = key

        for r, val in enumerate(value):
            if col==2:
                U_sum += val
            elif col==3:
                V_sum += val
            elif col==4:
                W_sum += val

            CURRENT_SHEET.cell(row=r+3, column=col).value = val        
    
        col +=1

    #Calculating averages.
    try:
        U_AVG = round(U_sum/ToT_CoUnT, 3)
        V_AVG = round(V_sum/ToT_CoUnT, 3)
        W_AVG = round(W_sum/ToT_CoUnT, 3)
    except ZeroDivisionError:
        print("No input data found!!\nDivision by zero occurred!")
        exit()

    #Inserting average values in output sheet.
    try:
        CURRENT_SHEET.cell(row=3, column=5).value = U_AVG
        CURRENT_SHEET.cell(row=3, column=6).value = V_AVG
        CURRENT_SHEET.cell(row=3, column=7).value = W_AVG
    except FileNotFoundError:
        print("Output file not found!!")
        exit()
    except ValueError:
        print("Row or column values must be at least 1 ")
        exit()


    #After getting input ready, here we will set the data with octants.
    PAIRING_Processed_Data_n_Octants(U_AVG, V_AVG, W_AVG, CURRENT_SHEET)


    CURRENT_SHEET.cell(row=1, column=14).value = "Overall Octant Count"

    CURRENT_SHEET.cell(row=1, column=24).value = "Rank #1 Should be highlighted YeLLoW"

    CURRENT_SHEET.cell(row=1, column=35).value = "Overall Transition Count"

    CURRENT_SHEET.cell(row=1, column=45).value = "Longest Subsequence Length"

    CURRENT_SHEET.cell(row=1, column=49).value = "Longest Subsequence Length with Range"
    
    CURRENT_SHEET.cell(row=2, column=36).value = "To"

    headers = ["T", "U", "V", "W", "U AVG", "V AVG", "W AVG","U'=U - U AVG", "V'=V - V AVG", "W'=W - W AVG", "Octant"]
    for i, header in enumerate(headers):
        CURRENT_SHEET.cell(row=2, column=i+1).value = header


    OvErAll_OcTanT_RaNk_CounT(CURRENT_SHEET, MoD, ToT_CoUnT)

    Set_MoDwise_COUNT(CURRENT_SHEET, MoD, ToT_CoUnT)

    Overall_Transition_Count(CURRENT_SHEET, ToT_CoUnT)

    #Function to add MoD wise count of transition.
    MoDwiSe_OvEraLl_TraNsiTioN_CoUNt(CURRENT_SHEET, MoD, ToT_CoUnT)

    Longest_SubSeq(CURRENT_SHEET, ToT_CoUnT)

    outputFile.save(outputFileName)


    data = CURRENT_SHEET.values
    columns = next(data)[0:]


    DatAFraMe = pandas.DataFrame(data, columns=columns)
    return outputFileName

#Setting count to 0 for octants.
def RsT_CnT(count):
    for item in OcT_SigN:
        count[item] = 0

#Method to initialise dictionary with 0 for "OcT_SigN" except 'left'.
def RsT_CnT_except(count, left):
    for item in OcT_SigN:
        if(item!=left):
            count[item] = 0



def Longest_SubSeq(CURRENT_SHEET, ToT_CoUnT):
	#This Dictinary will store consecutive sequence count.
    count = {}

    #This Dictionary wil store longest count.
    longest = {}

    #Initialing dictionary to 0 for all labels.
    RsT_CnT(count)
    RsT_CnT(longest)

    #Variable to check end value.
    end = -10

    #Iterating complete excel sheet.
    for i in range(0, ToT_CoUnT):
        currRow = i+3
        try:
            curr = int(CURRENT_SHEET.cell(column=11, row=currRow).value)

            #Comparing current and end value.
            if(curr==end):
                count[curr]+=1
                longest[curr] = max(longest[curr], count[curr])
                RsT_CnT_except(count, curr)
            else:
                count[curr]=1
                longest[curr] = max(longest[curr], count[curr])
                RsT_CnT_except(count, curr)
        except FileNotFoundError:
            print("File not found!!")
            exit()

        #Updating "end" variable.
        end = curr

    #Method to Count longest subsequence freq.
    Longest_SubSeq_freq(longest, CURRENT_SHEET, ToT_CoUnT)


def Longest_SubSeq_freq(longest, CURRENT_SHEET, ToT_CoUnT):
    #Dictinary to store consecutive sequence count.
    count = {}

    #Dictinary to store freq count.
    freq = {}

    #Dictionary to store time range.
    TiMeRaNgE = {}

    for label in OcT_SigN:
        TiMeRaNgE[label] = []

    #Initialing dictionary to 0 for all labels.
    RsT_CnT(count)
    RsT_CnT(freq)

    #Variable to check end value.
    end = -10

    #Iterating complete excel sheet.
    for i in range(0, ToT_CoUnT):
        currRow = i+3
        try:
            curr = int(CURRENT_SHEET.cell(column=11, row=currRow).value)
            
            #Comparing current and end value.
            if(curr==end):
                count[curr]+=1
            else:
                count[curr]=1        
                RsT_CnT_except(count, curr)

            #Upading freq.
            if(count[curr]==longest[curr]):
                freq[curr]+=1

                #Counting starting and ending time of longest subsequence.
                end = float(CURRENT_SHEET.cell(row=currRow, column=1).value)
                start = 100*end - longest[curr]+1
                start/=100

                #Inserting time interval into map.
                TiMeRaNgE[curr].append([start, end])

                #Reseting count dictionary.
                RsT_CnT(count)
            else:
                RsT_CnT_except(count, curr)
        except FileNotFoundError:
            print("File not found!")
            exit()
        except ValueError:
            print("File content is invalid!")
            exit()

        #Updating 'end' variable.
        end = curr

    #Setting freq table into sheet.
    SET_FREQ(longest, freq, CURRENT_SHEET)

    #Setting time range for longest subsequence.
    LoNgeSt_SubSeq_TiMe(longest, freq, TiMeRaNgE, CURRENT_SHEET)

#Method to set freq count to sheet.
def SET_FREQ(longest, freq, CURRENT_SHEET):
    #Iterating "OcT_SigN" and updating sheet.
    for i in range(9):
        for j in range(3):
            CURRENT_SHEET.cell(row = 3+i, column = 45+j).border = Blc_BORDER

    CURRENT_SHEET.cell(row=3, column=45).value= "Octant ##"
    CURRENT_SHEET.cell(row=3, column=46).value= "Longest Subsquence Length"
    CURRENT_SHEET.cell(row=3, column=47).value= "Count"

    for i, label in enumerate(OcT_SigN):
        currRow = i+3
        try:
            CURRENT_SHEET.cell(row=currRow+1, column=45).value = label	
            CURRENT_SHEET.cell(column=46, row=currRow+1).value = longest[label]
            CURRENT_SHEET.cell(column=47, row=currRow+1).value = freq[label]
        except FileNotFoundError:
            print("File not found!!")
            exit()

#Method to set time range for longest subsequence.
def LoNgeSt_SubSeq_TiMe(longest, freq, TiMeRaNgE, CURRENT_SHEET):
    #Naming columns number.
    lengthCol = 50
    freqCol = 51
    
    #Initial row, just after the header row.
    row = 4

    CURRENT_SHEET.cell(row=3, column = 49).value = "Octant ###"
    CURRENT_SHEET.cell(row=3, column = 50).value = "Longest Subsquence Length"
    CURRENT_SHEET.cell(row=3, column = 51).value = "Count"

    #Iterating all octants. 
    for octant in OcT_SigN:
        try:
            #Setting octant's longest subsequence and freq data.
            CURRENT_SHEET.cell(column=49, row=row).value = octant
            CURRENT_SHEET.cell(column=lengthCol, row=row).value = longest[octant]
            CURRENT_SHEET.cell(column=freqCol, row=row).value = freq[octant]

        except FileNotFoundError:
            print("File not found!")
            exit()

        row+=1

        try:
            #Setting default labels.
            CURRENT_SHEET.cell(column=49, row=row).value = "Time"
            CURRENT_SHEET.cell(column=lengthCol, row=row).value = "From"
            CURRENT_SHEET.cell(column=freqCol, row=row).value = "To"

        except FileNotFoundError:
            print("File not found!")
            exit()

        row+=1

        #Iterating time range values for each octants.
        for TiMeDaTa in TiMeRaNgE[octant]:
            try:
                #Setting time interval value.
                CURRENT_SHEET.cell(row=row, column=lengthCol).value = TiMeDaTa[0]
                CURRENT_SHEET.cell(row=row, column=freqCol).value = TiMeDaTa[1]
            except FileNotFoundError:
                print("File not found!!")
                exit()
            row += 1

    for i in range(3, row):
        for j in range(49, 52):
            CURRENT_SHEET.cell(row=i, column = j).border = Blc_BORDER



def MoDwiSe_OvEraLl_TraNsiTioN_CoUNt(CURRENT_SHEET, MoD, ToT_CoUnT):
    #Counting partitions w.r.t. MoD.
    try:
        totalPartition = ToT_CoUnT//MoD
    except ZeroDivisionError:
        print("MoD can't have 0 value")
        exit()

    #Checking MoD value range.
    if(MoD<0):
        raise Exception("MoD value should be in range of 1-30000")

    if(ToT_CoUnT%MoD!=0):
        totalPartition +=1

    #Initialising row start for data filling.
    RoW_StArT = 16

    #Iterating all partitions.
    for i in range (0,totalPartition):
        #Initialising start and end values.
        start = i*MoD
        end = min((i+1)*MoD-1, ToT_CoUnT-1)

        #Setting start-end values.
        try:
            CURRENT_SHEET.cell(column=35, row=RoW_StArT-1 + 13*i).value = "MoD Transition Count"
            CURRENT_SHEET.cell(column=35, row=RoW_StArT + 13*i).value = str(start) + "-" + str(end)
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()


        #Initialising empty dictionary.
        TRANScount = {}
        for a in range (1,5):
            for b in range(1,5):
                TRANScount[str(a)+str(b)]=0
                TRANScount[str(a)+str(-b)]=0
                TRANScount[str(-a)+str(b)]=0
                TRANScount[str(-a)+str(-b)]=0
                
        #Counting transition for range [start, end].
        for a in range(start, end+1):
            try:
                curr = CURRENT_SHEET.cell(column=11, row=a+3).value
                next = CURRENT_SHEET.cell(column=11, row=a+4).value
            except FileNotFoundError:
                print("Output file not found!!")
                exit()
            except ValueError:
                print("Row or column values must be at least 1 ")
                exit()

            #Incrementing count for within range value.
            if(next!=None):
                TRANScount[str(curr) + str(next)]+=1

        #Setting transition counts.
        Set_Transition_Count(RoW_StArT + 13*i, TRANScount, CURRENT_SHEET)



def Overall_Transition_Count(CURRENT_SHEET, ToT_CoUnT):
    #Setting value.

    #Initialising empty dictionary.
    TRANScount = {}
    for i in range (1,5):
        for j in range(1,5):
            TRANScount[str(i)+str(j)]=0
            TRANScount[str(i)+str(-j)]=0
            TRANScount[str(-i)+str(j)]=0
            TRANScount[str(-i)+str(-j)]=0
        
    #Iterating octants values to fill dictionary.
    start = 0

    #try and except block for string to int conversion.
    try:
        end = int(CURRENT_SHEET["K3"].value)
    except ValueError:
        print("Sheet input can't be CoNvErTed to int")
        exit()
    except TypeError:
        print("Sheet doesn't contain integer octant")
        exit()


    while(start<ToT_CoUnT-1):
        #try and except block for string to int conversion.
        try:
            curr = int(CURRENT_SHEET.cell(row= start+4, column=11).value)
            TRANScount[str(end) + str(curr)]+=1
            end = curr
        except ValueError:
            print("Sheet input can't be CoNvErTed to int")
            exit()
        except TypeError:
            print("Sheet doesn't contain integer octant")
            exit()

        start += 1
    
    #Setting transitions counted into sheet.
    Set_Transition_Count(2, TRANScount, CURRENT_SHEET)

#Function to set Transition count.
def Set_Transition_Count(row, TRANScount, CURRENT_SHEET):
    #Setting hard coded inputs.
    try:
        CURRENT_SHEET.cell(row=row, column=36).value = "To"
        CURRENT_SHEET.cell(row=row+1, column=35).value = "Octant #"
        CURRENT_SHEET.cell(row=row+2, column=34).value = "From"

        for i in range(35, 44):
            for j in range(row+1, row+1+9):
                CURRENT_SHEET.cell(row=j, column = i).border = Blc_BORDER


    except FileNotFoundError:
        print("Output file not found!!")
        exit()
    except ValueError:
        print("Row or column values must be at least 1 ")
        exit()

    #Setting Labels.
    for i, label in enumerate(OcT_SigN):
        try:
            CURRENT_SHEET.cell(row=row+1, column=i+36).value=label
            CURRENT_SHEET.cell(row=row+i+2, column=35).value=label
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

    #Setting data.
    for i, L1 in enumerate(OcT_SigN):
        maxi = -1

        for j, L2 in enumerate(OcT_SigN):
            val = TRANScount[str(L1)+str(L2)]
            maxi = max(maxi, val)

        

        for j, L2 in enumerate(OcT_SigN):
            try:
                CURRENT_SHEET.cell(row=row+i+2, column=36+j).value = TRANScount[str(L1)+str(L2)]
                if TRANScount[str(L1)+str(L2)] == maxi:
                    maxi = -1
                    CURRENT_SHEET.cell(row=row+i+2, column=36+j).fill = YeLLoW_bG
            except FileNotFoundError:
                print("Output file not found!")
                exit()
            except ValueError:
                print("Row or column values must be at least 1 ")
                exit()



def Set_MoDwise_COUNT(CURRENT_SHEET, MoD, ToT_CoUnT):
	#Initialising empty dictionary.
    count = {-1:0, 1:0, -2:0, 2:0, -3:0, 3:0, -4:0, 4:0}

    #Variable to store end row.
    lastRow = -1

    #Iterating loop to set count dictionary.
    start = 0
    while(start<ToT_CoUnT):
        try:
            count[int(CURRENT_SHEET.cell(row=start+3, column=11).value)] +=1
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

        start+=1
    
        try:    
            if(start%MoD==0):
                #Setting row data.
                try:
                    row = 4 + start//MoD
                    lastRow = row
                    CURRENT_SHEET.cell(row=row, column=14).value = str(start-MoD) + "-" + str(min(ToT_CoUnT, start-1))

                    for i, label in enumerate(OcT_SigN):
                        CURRENT_SHEET.cell(row=row, column=15+i).value = count[label]

                    SET_RANK_Count(row,count, CURRENT_SHEET)
                except FileNotFoundError:
                    print("Output file not found!!")
                    exit()
                except ValueError:
                    print("Row or column values must be at least 1 ")
                    exit()

                #Reset count values.
                count = {-1:0, 1:0,  -2:0, 2:0, -3:0, 3:0, -4:0, 4:0}
        except ZeroDivisionError:
            print("MoD can't have 0 value")
            exit()

    try:
        if(start%MoD!=0):
            #Setting row data.
            try:
                row = 5 + start//MoD
                lastRow = row
                CURRENT_SHEET.cell(row=row, column=14).value = str(start-MoD) + "-" + str(min(ToT_CoUnT, start-1))
                for i, label in enumerate(OcT_SigN):
                    CURRENT_SHEET.cell(row=row, column=15+i).value = count[label]
                
                SET_RANK_Count(row,count, CURRENT_SHEET)
            except FileNotFoundError:
                print("Output file not found!!")
                exit()
            except ValueError:
                print("Row or column values must be at least 1 ")
                exit()

    except ZeroDivisionError:
        print("MoD can't have 0 value")
        exit()

    if(lastRow!=-1):
        setOverallOctantRankMap(lastRow, CURRENT_SHEET)

def setOverallOctantRankMap(lastRow, CURRENT_SHEET):
    count = {-1:0, 1:0,  -2:0, 2:0, -3:0, 3:0, -4:0, 4:0}
    
    row =4
    while CURRENT_SHEET.cell(row=row, column=29).value is not None:
        oct = int(CURRENT_SHEET.cell(row=row, column=31).value)
        count[oct]+=1
        row+=1


    for i in range(9):
        for j in range(3):
            row = lastRow+2+i
            col = 29+j
            CURRENT_SHEET.cell(row=row, column = col).border = Blc_BORDER


    CURRENT_SHEET.cell(column=29, row=lastRow+2).value = "Octant ID"
    CURRENT_SHEET.cell(column=30, row=lastRow+2).value = "Octant Name "
    CURRENT_SHEET.cell(column=31, row=lastRow+2).value = "Count of Rank 1 MoD Values"

    for j, oct in enumerate(OcT_SigN):
        CURRENT_SHEET.cell(column=29, row=lastRow+3+j).value = oct
        CURRENT_SHEET.cell(column=30, row=lastRow+3+j).value = OcT_NaMe_Id_MaPPinG[oct]
        CURRENT_SHEET.cell(column=31, row=lastRow+3+j).value = count[oct]




def OvErAll_OcTanT_RaNk_CounT(CURRENT_SHEET, MoD, ToT_CoUnT):
    FiRsTRoW = ["Octant ID",1,-1,2,-2,3,-3,+4,-4,"Rank Octant 1", "Rank Octant -1","Rank Octant 2","Rank Octant -2","Rank Octant 3","Rank Octant -3","Rank Octant 4","Rank Octant -4","Rank1 Octant ID","Rank1 Octant Name"]

    ToTaLRoWs = ToT_CoUnT//MoD+1+1 
    if ToT_CoUnT%MoD!=0:
        ToTaLRoWs+=1


    for i, header in enumerate(FiRsTRoW):
        for j in range(ToTaLRoWs):
            CURRENT_SHEET.cell(row=3+j, column = 14+i).border = Blc_BORDER


    for i, header in enumerate(FiRsTRoW):
        CURRENT_SHEET.cell(row=3, column = i+14).value = header

    CURRENT_SHEET.cell(row=4, column = 13).value = "MoD " + str(MoD)

    SeTOveRallCounT(ToT_CoUnT, CURRENT_SHEET)

def SeTOveRallCounT(ToT_CoUnT, CURRENT_SHEET):	
	#Initialising count dictionary.
    count = {-1:0, 1:0, -2:0, 2:0, -3:0, 3:0, -4:0, 4:0}

    #Incrementing count dictionary data.
    try:
        for i in range (3,ToT_CoUnT+3):
            count[int(CURRENT_SHEET.cell(column=11, row=i).value)] = count[int(CURRENT_SHEET.cell(column=11, row=i).value)] +1
    except FileNotFoundError:
        print("Output file not found!!")
        exit()
    except ValueError:
        print("Sheet input can't be CoNvErTed to int or row/colum should be atleast 1")
        exit()
    except TypeError:
        print("Sheet doesn't contact valid octant value!!")
        exit()


    #Setting data into sheet.
    for i, label in enumerate(OcT_SigN):
        try:
            CURRENT_SHEET.cell(row=4, column=i+15).value = count[label]
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()
    SET_RANK_Count(4, count, CURRENT_SHEET)

def SET_RANK_Count(row,countMap, CURRENT_SHEET):
    #Copying the count list to sort.
    sortedCount = []
    count = []
    for label in OcT_SigN:
        count.append(countMap[label])

    for ct in count:
        sortedCount.append(ct)

    sortedCount.sort(reverse=True)
    rank = []

    for i, el in enumerate(count):
        for j, ell in enumerate(sortedCount):
            if(ell==el):
                rank.append(j+1)
                sortedCount[j] = -1
                break

    rank1Oct = -10

    for j in range(0,8):
        CURRENT_SHEET.cell(row = row, column=23+j).value = rank[j]
        if(rank[j]==1):
            rank1Oct = OcT_SigN[j]
            CURRENT_SHEET.cell(row = row, column=23+j).fill = YeLLoW_bG   

    OcT_NaMe_Id_MaPPinG = {1:"Internal outward interaction", -1:"External outward interaction", 2:"External Ejection", -2:"Internal Ejection", 3:"External inward interaction", -3:"Internal inward interaction", 4:"Internal sweep", -4:"External sweep"}
    CURRENT_SHEET.cell(row=row , column=31).value = rank1Oct
    CURRENT_SHEET.cell(row=row , column=32).value = OcT_NaMe_Id_MaPPinG[rank1Oct]

def PAIRING_Processed_Data_n_Octants(U_AVG, V_AVG, W_AVG, CURRENT_SHEET):
    start = 3
    time = CURRENT_SHEET.cell(start, 1).value

    #Iterating throught sheet.
    while(time!=None):
        #Calculating processed data.
        try:
            u1 = CURRENT_SHEET.cell(start, 2).value - U_AVG
            v1 = CURRENT_SHEET.cell(start, 3).value - V_AVG
            w1 = CURRENT_SHEET.cell(start, 4).value - W_AVG
            
            u1 = round(u1,3)
            v1 = round(v1,3)
            w1 = round(w1,3)

            oct = get_octant(u1, v1, w1)
        except FileNotFoundError:
            print("Input file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()


        #Setting processed data.
        try:
            CURRENT_SHEET.cell(row=start, column=8).value = u1
            CURRENT_SHEET.cell(row=start, column=9).value = v1
            CURRENT_SHEET.cell(row=start, column=10).value = w1
            CURRENT_SHEET.cell(row=start, column=11).value = oct
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()


        start = start+1
        try:
            time = CURRENT_SHEET.cell(start, 1).value
        except FileNotFoundError:
            print("Input file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

#Method based on if-else to return octant type.
def get_octant(x,y,z):
    if(x>=0 and y>=0):
        if(z>=0):
            return 1
        else:
            return -1
    if(x<0 and y>=0):
        if(z>=0):
            return 2
        else:
            return -2

    if(x<0 and y<0):
        if(z>=0):
            return 3
        else:
            return -3

    if(x>=0 and y<0):
        if(z>=0):
            return 4
        else:
            return -4
        
#Finally calling fuction for interface.   
StReAmLiT_PrOj2()
#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))