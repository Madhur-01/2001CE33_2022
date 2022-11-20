#MADHUR GARG, 2001CE33

#importing numpy and pandas
import numpy as np
import pandas as pd
import openpyxl


#reading the input file
import os
path = r"input"
path2 = r"output"
for file in os.listdir(path):
    df = pd.read_excel(os.path.join(path,file))

    #data preprocessing
    
    df.at[0,'U_avg']  = round(df['U'].mean(),3)
    df.at[0,'V_avg']  = round(df['V'].mean(),3)
    df.at[0,'W_avg']  = round(df['W'].mean(),3)


    df["U'"] = round(df['U'] - df.at[0,'U_avg'],3)
    df["V'"] = round(df['V'] - df.at[0,'V_avg'],3)
    df["W'"] = round(df['W'] - df.at[0,'W_avg'],3)


    #definig a function to categorise data in different octant
    def octant(x,y,z) :
        if x>0 :
            if y>0:
                if z>0 :
                    return +1
                else :
                    return -1
            else :
                if z>0 :
                    return +4
                else :
                    return -4
        else :
            if y> 0 :
                if z>0:
                    return +2
                else :
                    return -2
            else :
                if z>0:
                    return +3
                else:
                    return -3

    #applying the above function           
    df['octant']         =   df.apply([lambda x : octant(x["U'"],x["V'"],x["W'"])], axis=1)

    #leaving an empty column
    df.at[1,''] = 'User Input'

    #counting individual octant uing value_counts function
    df.at[0,'Octant ID'] =   'Overall Count'
    df.at[0,'1']        =   str(df['octant'].value_counts()[+1])
    df.at[0,'-1']        =   str(df['octant'].value_counts()[-1])
    df.at[0,'2']        =   str(df['octant'].value_counts()[+2])
    df.at[0,'-2']        =   str(df['octant'].value_counts()[-2])
    df.at[0,'3']        =   str(df['octant'].value_counts()[+3])
    df.at[0,'-3']        =   str(df['octant'].value_counts()[-3])
    df.at[0,'4']        =   str(df['octant'].value_counts()[+4])
    df.at[0,'-4']        =   str(df['octant'].value_counts()[-4])


    #asking user for input
    mod = int(input('enter the value of mod: '))


    df.at[1,'Octant ID'] = 'Mod '+ str(mod)


    size = len(df['octant'])
    m=0
    #using a while loop to split the data 
    while(size>0):
        temp = mod
        if m == 0: #starting from value 0
            x = 0
        else:
            x = m*temp + 1 

        y = m*temp+mod
        if size<mod:
            mod = size
            size = 0
    
    
        #inserting range and their corresponding data
        m1 = str(x)
        m2= str(y)
        df.at[m+2,'Octant ID'] = m1 +'-'+m2 
        df2 = df.loc[x:y] 
   
        df.at[m+2,'1'] = str(df2['octant'].value_counts()[+1])
        df.at[m+2,'2'] = str(df2['octant'].value_counts()[+2])
        df.at[m+2,'3'] = str(df2['octant'].value_counts()[+3])
        df.at[m+2,'4'] = str(df2['octant'].value_counts()[+4])
        df.at[m+2,'-1'] = str(df2['octant'].value_counts()[-1])
        df.at[m+2,'-2'] = str(df2['octant'].value_counts()[-2])
        df.at[m+2,'-3'] = str(df2['octant'].value_counts()[-3])
        df.at[m+2,'-4'] = str(df2['octant'].value_counts()[-4])
    

        m = m + 1
        size = size - mod
    
    
    #defining a dictionary of name of octants
    dict = {'1':'Internal outward interaction','-1':'External outward interaction','2':'External Ejection','-2':'Internal Ejection','3':'External inward interaction','-3':'Internal inward interaction','4':'Internal sweep','-4':'External sweep'}

    #making a list of counts in ascending order
    list = []
    for i in range(-4,0):
        list.append(df.at[0,str(i)])
    for i in range(1,5):
        list.append(df.at[0,str(i)])
    list.sort()

    #filling the rank of counts
    k = 8
    for i in range(len(list)):
        df.at[0,'rank'+(df == list[i]).idxmax(axis=1)[0]] = k
        k = k-1
    #filling the octant with highest count
    df.at[0,'Rank1_Octant_ID'] = int((df == list[7]).idxmax(axis=1)[0])
    #filling the name of octant
    df.at[0,'Rank1 Octant Name'] = dict[(df == list[7]).idxmax(axis=1)[0]]


    #making and filling list of counts in ascending order
    for j in range(2,int(len(df)/mod)+2):
      
        list = []
        for i in range(-4,0):
            list.append(df.at[j,str(i)])
        for i in range(1,5):
            list.append(df.at[j,str(i)])
        list.sort()
        k = 8
        for i in range(len(list)):
            df.at[j,'rank'+(df == list[i]).idxmax(axis=1)[j]] = k
            k = k-1
        #filling the octant with highest count
        df.at[j,'Rank1_Octant_ID'] = int((df == list[7]).idxmax(axis=1)[j])
        #filling the name of octant
        df.at[j,'Rank1 Octant Name'] = dict[(df == list[7]).idxmax(axis=1)[j]]


    #calculating count of rank1 mod values
    df.at[6+int(len(df)/mod),'-3'] = 'Octant ID'
    df.at[6+int(len(df)/mod),'4'] = 'Octant Name'
    df.at[6+int(len(df)/mod),'-4'] = 'Count of rank1 mod values'
    k = 1
    for i in range(-4,0):
        df.at[k+6+int(len(df)/mod),'-3'] = i
        df.at[k+6+int(len(df)/mod),'4'] = dict[str(i)]
        try :
          df.at[k+6+int(len(df)/mod),'-4'] = df['Rank1_Octant_ID'].value_counts()[int(i)]
        except :
            df.at[k+6+int(len(df)/mod),'-4'] = 0
        k = k+1
    for i in range(1,5):
        df.at[k+6+int(len(df)/mod),'-3'] = i
        df.at[k+6+int(len(df)/mod),'4'] = dict[str(i)]
        try :
          df.at[k+6+int(len(df)/mod),'-4'] = df['Rank1_Octant_ID'].value_counts()[int(i)]
        except :
            df.at[k+6+int(len(df)/mod),'-4'] = 0
        k = k+1


    #leaving an empty column
    df.at[0,'blank'] = ''

        #defining a function to get transition count 
    def transition_count(df,l,m):
        k=0
        for i in range(len(df)-1):
            if df.at[i,'octant'] == l and df.at[i+1,'octant'] ==m:
                k = k+1
        return k


    s = int(len(df)/mod)
    df.at[1,'Overall Transition Count'] =  'From'
    df.at[1,'(1)']=  'to'
    df.at[2,'(1)']=  1
    df.at[2,'(-1)']= -1
    df.at[2,'(2)']=  2
    df.at[2,'(-2)']= -2
    df.at[2,'(3)']=  3
    df.at[2,'(-3)']= -3
    df.at[2,'(4)']=  4
    df.at[2,'(-4)']= -4

    #df.at[s+9,'']= 'From'
    df.at[2,'Overall Transition Count'] = "Octant"  
    df.at[3,'Overall Transition Count']=  -4
    df.at[4,'Overall Transition Count']= -3
    df.at[5,'Overall Transition Count']= -2
    df.at[6,'Overall Transition Count']= -1
    df.at[7,'Overall Transition Count']=  1
    df.at[8,'Overall Transition Count']=  2
    df.at[9,'Overall Transition Count']=  3
    df.at[10,'Overall Transition Count']=  4

    #calculating overall transition count
    for x in range(3,7):
        for y in range(-4,5) :
            df.at[x,"(" +str(y)+ ")"] = transition_count(df,x-7,y)
    for x in range(7,11):
        for y in range(-4,5) :
            df.at[x,"("+str(y)+")"] = transition_count(df,x-6,y)
            
            
    size = len(df['octant'])
    q=1

    #defining a function for mod transition count
    def mod_transition_count(df,mod,l,m):
            k=0
            if mod*q-1<len(df):
                for i in range(mod*(q-1),mod*q-1):
                    if df.at[i,'octant'] == l and df.at[i+1,'octant'] ==m:
                        k = k+1
            else:
                for i in range(mod*(q-1),len(df)-1):
                    if df.at[i,'octant'] == l and df.at[i+1,'octant'] ==m:
                        k = k+1
            return k 
    
    mod = 5000
    #using a while loop to calculate mod transition count
    while(size>0):
   

        if size<mod:
            size=0
       
        df.at[14*q,'Overall Transition Count'] = 'Mod Transition Count'
        df.at[1+14*q,'Overall Transition Count'] =  'to'
        df.at[2+14*q,'(1)']=  1
        df.at[2+14*q,'(-1)']= -1
        df.at[2+14*q,'(2)']=   2
        df.at[2+14*q,'(-2)']= -2
        df.at[2+14*q,'(3)']=   3
        df.at[2+14*q,'(-3)']= -3
        df.at[2+14*q,'(4)']=   4
        df.at[2+14*q,'(-4)']= -4

        df.at[2+14*q,'']= str(mod*(q-1))+"-"+str(mod*q)
        df.at[3+14*q,'']= 'From'
        df.at[2+14*q,'Overall Transition Count'] = "Count"  
        df.at[3+14*q,'Overall Transition Count'] =  -4
        df.at[4+14*q,'Overall Transition Count'] =  -3
        df.at[5+14*q,'Overall Transition Count'] =  -2
        df.at[6+14*q,'Overall Transition Count']  =  -1
        df.at[7+14*q,'Overall Transition Count'] =   1
        df.at[8+14*q,'Overall Transition Count'] =   2
        df.at[9+14*q,'Overall Transition Count'] =   3
        df.at[10+14*q,'Overall Transition Count'] =   4

        #calculating overall transition count
        for x in range(4+14*q,8+14*q):
            for y in range(-4,5) :
                df.at[x,"("+str(y)+")"] = mod_transition_count(df,mod,x-8-14*q,y)
        for x in range(8+14*q,12+14*q):
            for y in range(-4,5) :
                df.at[x,"("+str(y)+")"] = mod_transition_count(df,mod,x-7-14*q,y)


        q = q + 1
        size = size - mod
    #deleting extra column
    del df['(0)']


    #leaving an empty column
    df.at[0,'blank2'] = ''


    #making columns for subsequence
    df['Octant_No'] = ''
    df['Longest_Subsequence_Length'] = ''
    df['Count'] = ''

    l = [1,-1,2,-2,3,-3,4,-4] #making a list of all the octants
    l1 = df['octant'].tolist()
    i=0
    for x in l: #finding subsequence for every octant
        df.at[i,'Octant_No'] = x
        count = 1
        temp = 1
        mx = 0
        for y in range(len(l1)-1):
            if x == l1[y] and x == l1[y+1]:
                temp += 1
            else:
                if mx == temp:
                    count += 1
                elif mx < temp:
                    count = 1
                mx = max(mx,temp)
                temp = 1
        df.at[i,'Longest_Subsequence_Length'] = mx
        df.at[i,'Count'] = count
        i += 1
    df.at[1,'blank3'] = ''
    i=0
    j=0
    df[''] = ''
    df['Octant_No.'] = ''
    df['Longest_Subsequence_Length_'] = ''
    df['count'] = ''
    for x in l: #finding subsequence for every octant
        t1 = df.at[i,'Longest_Subsequence_Length']
        df.at[j,'Octant_No.'] = df.at[i,'Octant_No']
        df.at[j,'Longest_Subsequence_Length_'] = df.at[i,'Longest_Subsequence_Length']
        df.at[j,'count'] = df.at[i,'Count']
        j += 1
        df.at[j,'Octant_No.'] = 'T'
        df.at[j,'Longest_Subsequence_Length_'] = 'From'
        df.at[j,'count'] = 'To'
        j += 1
        temp = 1
        for y in range(len(l1)-1):
            if x == l1[y] and x == l1[y+1]:
                temp += 1
            elif temp == df.at[i,'Longest_Subsequence_Length']:
                df.at[j,'Longest_Subsequence_Length_'] = df.at[y-temp+1,'T']
                df.at[j,'count'] = df.at[y,'T']
                j += 1
                temp = 1
            else:
                temp = 1
        i += 1
    df.to_excel('output/'+file+'_output.xlsx',index = False)
    

    from openpyxl.styles import PatternFill   
    wb = openpyxl.load_workbook(os.path2.join(path2,file))

    # entries = os.listdir('input/')
    # from collections import OrderedDict
    # import numpy as np
    # #Help
    # from openpyxl.styles.borders import Border, Side
    # from openpyxl.styles import PatternFill
    # from openpyxl import Workbook
    # fill_cell = PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid')

    # def octant_analysis(mod=5000):
	#     for file1 in entries:  
    # sheet_obj.cell(row=x, column=y).border = thin_border       
    print(df)

